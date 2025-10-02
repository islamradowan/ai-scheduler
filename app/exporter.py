import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import os
from datetime import datetime

def get_section_from_student_id(student_id, students_df=None):
    """Get section from student ID by looking up in students_df"""
    if not student_id or students_df is None:
        return ""
    
    student_id_str = str(student_id)
    
    try:
        # Convert students_df student_id to string for comparison
        students_df_copy = students_df.copy()
        students_df_copy['student_id'] = students_df_copy['student_id'].astype(str)
        
        # Find matching student
        student_row = students_df_copy[students_df_copy['student_id'] == student_id_str]
        
        if not student_row.empty:
            section = student_row.iloc[0].get('section', '')
            return str(section)
    except Exception as e:
        pass
    
    return ""

def get_batch_from_student_id(student_id, students_df=None):
    """Extract batch from student ID based on batch_type (first 4 for regular, first 5 for evening)"""
    if not student_id:
        print(f"Debug: Empty student_id")
        return ""
    
    if students_df is None:
        print(f"Debug: students_df is None")
        return ""
    
    # Convert to string and get digits only
    student_id_str = str(student_id)
    digits_only = ''.join(filter(str.isdigit, student_id_str))
    
    print(f"Debug: student_id={student_id_str}, digits_only={digits_only}")
    print(f"Debug: students_df shape: {students_df.shape}")
    print(f"Debug: students_df columns: {list(students_df.columns)}")
    print(f"Debug: First few student_ids in df: {students_df['student_id'].head().tolist()}")
    
    if not digits_only:
        return ""
    
    # Look up student in students_df to get batch_type
    try:
        # Convert students_df student_id to string for comparison
        students_df_copy = students_df.copy()
        students_df_copy['student_id'] = students_df_copy['student_id'].astype(str)
        
        print(f"Debug: Looking for student_id '{student_id_str}' in converted df")
        
        # Find matching student
        student_row = students_df_copy[students_df_copy['student_id'] == student_id_str]
        
        print(f"Debug: Found {len(student_row)} matching rows")
        
        if not student_row.empty:
            batch_type = student_row.iloc[0].get('batch_type', '').lower()
            print(f"Debug: Found batch_type='{batch_type}' for student {student_id_str}")
            
            # Extract digits based on batch_type
            if 'evening' in batch_type:
                result = digits_only[:5]  # Evening: first 5 digits
                print(f"Debug: Evening student, returning first 5 digits: {result}")
                return result
            else:
                result = digits_only[:4]  # Regular: first 4 digits
                print(f"Debug: Regular student, returning first 4 digits: {result}")
                return result
        else:
            print(f"Debug: No matching student found for {student_id_str}")
    except Exception as e:
        print(f"Debug: Exception in lookup: {e}")
    
    # Default fallback: assume regular (first 4 digits)
    result = digits_only[:4]
    print(f"Debug: Using fallback, returning: {result}")
    return result

def get_day_name(date_str):
    """Get day name from date string"""
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return date_obj.strftime('%A').lower()
    except:
        return ""

def format_date_time(date_str, time_str):
    """Format date and time for display"""
    try:
        # Convert date format
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        formatted_date = date_obj.strftime('%d.%m.%y')
        day_name = date_obj.strftime('%A').lower()
        
        # Convert time to 12-hour format
        time_12hr = convert_to_12hour_format(time_str)
        
        return f"{formatted_date} ({day_name})\ntime: {time_12hr}"
    except:
        return f"{date_str}\ntime: {time_str}"

def convert_to_12hour_format(time_str):
    """Convert 24-hour time format to 12-hour AM/PM format"""
    if not time_str or '-' not in time_str:
        return time_str
    
    try:
        start_time, end_time = time_str.split('-')
        
        # Convert start time
        start_hour, start_min = start_time.split(':')
        start_hour = int(start_hour)
        start_period = 'AM' if start_hour < 12 else 'PM'
        if start_hour == 0:
            start_hour = 12
        elif start_hour > 12:
            start_hour -= 12
        start_formatted = f"{start_hour}:{start_min} {start_period}"
        
        # Convert end time
        end_hour, end_min = end_time.split(':')
        end_hour = int(end_hour)
        end_period = 'AM' if end_hour < 12 else 'PM'
        if end_hour == 0:
            end_hour = 12
        elif end_hour > 12:
            end_hour -= 12
        end_formatted = f"{end_hour}:{end_min} {end_period}"
        
        return f"{start_formatted} - {end_formatted}"
    except:
        return time_str

def export_excel(timetable, out_path):
    """
    Export timetable to Excel with multiple sheets.
    
    Args:
        timetable: List of exam assignments
        out_path: Output file path
    """
    # Ensure outputs directory exists
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Timetable sheet
    timetable_data = []
    for exam in timetable:
        for assignment in exam.get('assignments', []):
            timetable_data.append({
                'Course ID': exam['course_id'],
                'Date': exam.get('slot_date', ''),
                'Time': exam.get('slot_time', ''),
                'Room': assignment.get('room_id', ''),
                'Students': len(assignment.get('students', [])),
                'Invigilator': assignment.get('invigilator', 'Unassigned'),
                'Status': exam.get('status', 'scheduled')
            })
    
    timetable_df = pd.DataFrame(timetable_data)
    ws1 = wb.create_sheet("Timetable")
    for r in dataframe_to_rows(timetable_df, index=False, header=True):
        ws1.append(r)
    
    # Rooms sheet
    rooms_data = []
    for exam in timetable:
        for assignment in exam.get('assignments', []):
            rooms_data.append({
                'Room ID': assignment.get('room_id', ''),
                'Course': exam['course_id'],
                'Date': exam.get('slot_date', ''),
                'Time': exam.get('slot_time', ''),
                'Capacity Used': len(assignment.get('students', []))
            })
    
    rooms_df = pd.DataFrame(rooms_data)
    ws2 = wb.create_sheet("Rooms")
    for r in dataframe_to_rows(rooms_df, index=False, header=True):
        ws2.append(r)
    
    # Seat Maps sheet
    seat_data = []
    for exam in timetable:
        for assignment in exam.get('assignments', []):
            for seat in assignment.get('seat_assignments', []):
                seat_data.append({
                    'Course': exam['course_id'],
                    'Room': assignment.get('room_id', ''),
                    'Student ID': seat.get('student_id', ''),
                    'Row': seat.get('row', ''),
                    'Column': seat.get('column', '')
                })
    
    seats_df = pd.DataFrame(seat_data)
    ws3 = wb.create_sheet("SeatMaps")
    for r in dataframe_to_rows(seats_df, index=False, header=True):
        ws3.append(r)
    
    # Invigilators sheet
    invig_data = []
    for exam in timetable:
        for assignment in exam.get('assignments', []):
            if assignment.get('invigilator'):
                invig_data.append({
                    'Teacher ID': assignment.get('invigilator', ''),
                    'Course': exam['course_id'],
                    'Room': assignment.get('room_id', ''),
                    'Date': exam.get('slot_date', ''),
                    'Time': exam.get('slot_time', ''),
                    'Load Score': assignment.get('load_balance_score', 0)
                })
    
    invig_df = pd.DataFrame(invig_data)
    ws4 = wb.create_sheet("Invigilators")
    for r in dataframe_to_rows(invig_df, index=False, header=True):
        ws4.append(r)
    
    wb.save(out_path)

def export_pdf(timetable, out_path, students_df=None):
    """
    Export timetable to PDF with summary and seat maps.
    
    Args:
        timetable: List of exam assignments
        out_path: Output file path
    """
    # Ensure outputs directory exists
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    
    doc = SimpleDocTemplate(out_path, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title = Paragraph("Exam Timetable Summary", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Group exams by date and time
    from collections import defaultdict
    
    exam_groups = defaultdict(list)
    
    for exam in timetable:
        total_students = sum(len(assignment.get('students', [])) for assignment in exam.get('assignments', []))
        
        # Only include exams with students
        if total_students > 0:
            date_time_key = f"{exam.get('slot_date', '')}_{exam.get('slot_time', '')}"
            exam_groups[date_time_key].append(exam)
    
    # Sort groups by date and time
    def sort_key(date_time_key):
        try:
            from datetime import datetime
            date_str, time_str = date_time_key.split('_')
            start_time = time_str.split('-')[0] if '-' in time_str else '00:00'
            datetime_str = f"{date_str} {start_time}"
            return datetime.strptime(datetime_str, '%Y-%m-%d %H:%M')
        except:
            return datetime.min
    
    sorted_groups = sorted(exam_groups.items(), key=lambda x: sort_key(x[0]))
    
    # Summary table with new format
    summary_data = [['Date and Time', 'Batch', 'Section', 'Course Code', 'Course Title']]
    
    for date_time_key, exams in sorted_groups:
        # Collect all data for this time slot
        all_batches = set()
        all_sections = set()
        course_codes = []
        course_titles = []
        
        for exam in exams:
            # Get all student IDs for this exam
            all_students = []
            for assignment in exam.get('assignments', []):
                all_students.extend(assignment.get('students', []))
            
            # Extract unique batches and sections from student IDs
            for student_id in all_students:
                batch = get_batch_from_student_id(student_id, students_df)
                if batch:
                    all_batches.add(batch)
                
                # Get section from students_df
                section = get_section_from_student_id(student_id, students_df)
                if section:
                    all_sections.add(section)
            
            # Collect course info
            course_code = exam.get('course_code', exam.get('course_id', ''))
            course_title = exam.get('course_name', exam.get('course_title', ''))
            
            if course_code:
                course_codes.append(course_code)
            if course_title:
                course_titles.append(course_title)
        
        # Format date and time (use first exam for date/time)
        first_exam = exams[0]
        date_time_formatted = format_date_time(
            first_exam.get('slot_date', ''), 
            first_exam.get('slot_time', '')
        )
        
        # Group batches and course codes in pairs
        def group_in_pairs(items):
            grouped = []
            for i in range(0, len(items), 2):
                pair = items[i:i+2]
                grouped.append(','.join(pair))
            return '\n'.join(grouped)
        
        # Combine multiple courses with line breaks
        sorted_batches = sorted(all_batches)
        sorted_sections = sorted(all_sections)
        combined_batches = group_in_pairs(sorted_batches)
        combined_sections = group_in_pairs(sorted_sections)
        combined_course_codes = group_in_pairs(course_codes)
        # Break long course titles and ensure each is on a separate line
        def format_course_title(title):
            title_str = str(title)
            # Break after "Microprocessor, Microcontrollers and " for long titles
            if "Microprocessor, Microcontrollers and " in title_str:
                return title_str.replace("Microprocessor, Microcontrollers and ", "Microprocessor, Microcontrollers and \n")
            return title_str
        
        if course_titles:
            formatted_titles = [format_course_title(title) + ',' for title in course_titles]
            combined_course_titles = '\n'.join(formatted_titles)
        else:
            combined_course_titles = ''
        
        summary_data.append([
            date_time_formatted,
            combined_batches,
            combined_sections,
            combined_course_codes,
            combined_course_titles
        ])
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Room seat plans organized by exam day
    story.append(Paragraph("Seat Plans", styles['Heading1']))
    story.append(Spacer(1, 12))
    
    # Group exams by date
    exams_by_date = defaultdict(list)
    
    for exam in timetable:
        if exam.get('assignments') and any(assignment.get('students') for assignment in exam.get('assignments', [])):
            exam_date = exam.get('slot_date', 'Unknown')
            exams_by_date[exam_date].append(exam)
    
    # Sort dates
    sorted_dates = sorted(exams_by_date.keys())
    
    # Create seat plans for each exam day
    for exam_date in sorted_dates:
        # Add date header
        date_obj = datetime.strptime(exam_date, '%Y-%m-%d')
        formatted_date = date_obj.strftime('%d.%m.%y (%A)')
        story.append(Paragraph(f"Exam Date: {formatted_date}", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Group assignments by room for this date
        room_assignments = defaultdict(list)
        
        for exam in exams_by_date[exam_date]:
            for assignment in exam.get('assignments', []):
                if assignment.get('students'):  # Only rooms with students
                    room_id = assignment.get('room_id', 'Unknown')
                    room_assignments[room_id].append({
                        'exam': exam,
                        'assignment': assignment
                    })
        
        # Create seat plan for each room on this date
        for room_id, assignments in room_assignments.items():
            # Collect all course codes for this room
            course_codes = []
            all_students = []
            
            for item in assignments:
                exam = item['exam']
                assignment = item['assignment']
                
                course_code = exam.get('course_code', exam.get('course_id', ''))
                if course_code:
                    course_codes.append(course_code)
                
                all_students.extend(assignment.get('students', []))
        
        # Determine number of columns (4-7 based on room size)
        num_students = len(all_students)
        if num_students <= 20:
            num_cols = 4
        elif num_students <= 40:
            num_cols = 5
        elif num_students <= 60:
            num_cols = 6
        else:
            num_cols = 7
        
        # Create room table
        seat_data = []
        
        # First row: Room ID
        seat_data.append([f"Room ID: {room_id}"])
        
        # Second row: Course codes (with line breaks after every 10 codes)
        def format_course_codes(codes):
            if not codes:
                return ""
            
            formatted_lines = []
            for i in range(0, len(codes), 8):
                line_codes = codes[i:i+8]
                formatted_lines.append(', '.join(line_codes))
            
            return '\n'.join(formatted_lines)
        
        course_codes_formatted = format_course_codes(course_codes)
        seat_data.append([f"Course Codes: {course_codes_formatted}"])
        
        # Third row: Column headers
        column_headers = [f"Column {i+1}" for i in range(num_cols)]
        seat_data.append(column_headers)
        
        # Add student IDs in columns (column-wise filling)
        num_rows = len(all_students) // num_cols + (1 if len(all_students) % num_cols > 0 else 0)
        
        for row in range(num_rows):
            row_data = []
            for col in range(num_cols):
                student_idx = col * num_rows + row
                if student_idx < len(all_students):
                    row_data.append(str(all_students[student_idx]))
                else:
                    row_data.append('')
            seat_data.append(row_data)
        
        # Create table
        room_table = Table(seat_data)
        room_table.setStyle(TableStyle([
            # Room ID row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('SPAN', (0, 0), (num_cols-1, 0)),  # Span room ID across all columns
            
            # Course codes row styling
            ('BACKGROUND', (0, 1), (-1, 1), colors.lightblue),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 10),
            ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
            ('SPAN', (0, 1), (num_cols-1, 1)),  # Span course codes across all columns
            
            # Column headers row styling
            ('BACKGROUND', (0, 2), (-1, 2), colors.grey),
            ('TEXTCOLOR', (0, 2), (-1, 2), colors.white),
            ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 2), (-1, 2), 9),
            ('ALIGN', (0, 2), (-1, 2), 'CENTER'),
            
            # Student ID rows styling
            ('BACKGROUND', (0, 3), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 3), (-1, -1), 8),
            ('ALIGN', (0, 3), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Grid for all cells
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(room_table)
        story.append(Spacer(1, 15))
        
        # Add extra space between exam dates
        story.append(Spacer(1, 30))
    
    doc.build(story)

if __name__ == "__main__":
    # Demo with sample data
    sample_timetable = [
        {
            'course_id': 'C001',
            'slot_date': '2024-05-01',
            'slot_time': '09:00-12:00',
            'status': 'scheduled',
            'assignments': [
                {
                    'room_id': 'R001',
                    'students': ['S001', 'S002', 'S003'],
                    'invigilator': 'T001',
                    'seat_assignments': [
                        {'student_id': 'S001', 'row': 1, 'column': 1},
                        {'student_id': 'S002', 'row': 1, 'column': 2},
                        {'student_id': 'S003', 'row': 2, 'column': 1}
                    ]
                }
            ]
        },
        {
            'course_id': 'C002',
            'slot_date': '2024-05-01',
            'slot_time': '14:00-17:00',
            'assignments': [
                {
                    'room_id': 'R002',
                    'students': ['S004', 'S005'],
                    'invigilator': 'T002',
                    'seat_assignments': [
                        {'student_id': 'S004', 'row': 1, 'column': 1},
                        {'student_id': 'S005', 'row': 1, 'column': 2}
                    ]
                }
            ]
        }
    ]
    
    # Export to Excel
    excel_path = "outputs/demo_timetable.xlsx"
    export_excel(sample_timetable, excel_path)
    print(f"Excel exported to: {excel_path}")
    
    # Export to PDF
    pdf_path = "outputs/demo_timetable.pdf"
    export_pdf(sample_timetable, pdf_path)
    print(f"PDF exported to: {pdf_path}")